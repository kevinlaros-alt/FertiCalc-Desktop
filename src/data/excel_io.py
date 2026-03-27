"""Excel I/O — lees/schrijf recepten en meststoffen naar/van .xlsx."""

import os
import sys
import json
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from ..engine.fertilizers import FERTILIZERS, FertilizerDef
from ..engine.calculations import TankConfig, WaterAnalysis, DrainAnalysis, TargetValues


DATA_FILE = 'FertiCalc_data.xlsx'
SHEET_FERTS = 'Meststoffen'
SHEET_RECIPES = 'Recepten'
SHEET_RECIPE_FERTS = 'ReceptMeststoffen'
SHEET_WATER = 'Wateranalyses'
SHEET_HISTORY = 'Historie'


def _get_path() -> str:
    """Get path to data file. Uses ~/Documents/FertiCalc/ for writable storage."""
    user_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'FertiCalc')
    user_path = os.path.join(user_dir, DATA_FILE)
    if os.path.exists(user_path):
        return user_path

    bundle_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    for base in [bundle_dir, os.path.dirname(os.path.dirname(bundle_dir)), os.getcwd()]:
        p = os.path.join(base, DATA_FILE)
        if os.path.exists(p):
            return p

    os.makedirs(user_dir, exist_ok=True)
    return user_path


def ensure_database() -> str:
    """Create database Excel if it doesn't exist. Returns path."""
    path = _get_path()
    if os.path.exists(path):
        return path

    wb = Workbook()

    # Meststoffen sheet
    ws = wb.active
    ws.title = SHEET_FERTS
    ws.append(['ID', 'Naam', 'Tank', 'Element1', 'Pct1', 'Element2', 'Pct2', 'EC Waarde', 'Molaire Massa', 'Eenheid', 'Vloeibaar', 'Dichtheid'])
    for f in FERTILIZERS:
        ai1 = f.active_ingredients[0] if len(f.active_ingredients) > 0 else None
        ai2 = f.active_ingredients[1] if len(f.active_ingredients) > 1 else None
        ws.append([
            f.id, f.name, f.tank,
            ai1.element if ai1 else '', ai1.percentage if ai1 else 0,
            ai2.element if ai2 else '', ai2.percentage if ai2 else 0,
            f.ec_value, f.molar_mass or '', f.unit,
            f.is_liquid, f.liquid_density if f.is_liquid else '',
        ])

    # Recepten sheet (extended with mode, drain, targets)
    ws2 = wb.create_sheet(SHEET_RECIPES)
    ws2.append(['ID', 'Naam', 'Gewas', 'Aangemaakt', 'Gewijzigd',
                'Tank A L', 'Tank A Verd', 'Tank B L', 'Tank B Verd',
                'Tank C L', 'Tank C Verd', 'Dosering', 'Gewenste EC',
                'Mode', 'Drain %', 'Drain Analysis JSON', 'Target Values JSON'])

    # ReceptMeststoffen sheet
    ws3 = wb.create_sheet(SHEET_RECIPE_FERTS)
    ws3.append(['Recept ID', 'Meststof ID', 'Hoeveelheid'])

    # Wateranalyses sheet (extended with Na)
    ws4 = wb.create_sheet(SHEET_WATER)
    ws4.append(['Recept ID', 'EC', 'NO3', 'K', 'Mg', 'Ca', 'P2O5', 'S', 'Cl', 'Na',
                'Fe', 'Mn', 'Zn', 'B', 'Cu', 'Mo'])

    # Historie sheet
    ws5 = wb.create_sheet(SHEET_HISTORY)
    ws5.append(['Recept ID', 'Datum', 'Snapshot'])

    wb.save(path)
    return path


def _next_recipe_id(ws) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_id = max(max_id, int(row[0]))
    return max_id + 1


def list_recipes() -> List[Dict]:
    path = ensure_database()
    wb = load_workbook(path, read_only=True)
    ws = wb[SHEET_RECIPES]
    recipes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        recipes.append({
            'id': int(row[0]),
            'name': row[1] or '',
            'crop': row[2] or '',
            'modified': str(row[4] or row[3] or ''),
        })
    wb.close()
    return recipes


def _safe_col(row, idx, default=None):
    """Safely get column value (handles old files with fewer columns)."""
    try:
        return row[idx] if row[idx] is not None else default
    except (IndexError, TypeError):
        return default


def load_recipe(recipe_id: int) -> Optional[Dict]:
    path = ensure_database()
    wb = load_workbook(path, read_only=True)

    ws = wb[SHEET_RECIPES]
    recipe = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and int(row[0]) == recipe_id:
            mode = _safe_col(row, 13, 'standard')
            drain_pct = float(_safe_col(row, 14, 0) or 0)
            drain_json = _safe_col(row, 15, '{}')
            target_json = _safe_col(row, 16, '{}')

            # Parse drain analysis
            drain = DrainAnalysis()
            if drain_json and drain_json != '{}':
                try:
                    d = json.loads(drain_json) if isinstance(drain_json, str) else {}
                    for k, v in d.items():
                        if hasattr(drain, k):
                            setattr(drain, k, float(v))
                except (json.JSONDecodeError, TypeError):
                    pass

            # Parse target values
            targets = TargetValues()
            if target_json and target_json != '{}':
                try:
                    t = json.loads(target_json) if isinstance(target_json, str) else {}
                    for k, v in t.items():
                        if hasattr(targets, k):
                            setattr(targets, k, float(v))
                except (json.JSONDecodeError, TypeError):
                    pass

            recipe = {
                'id': int(row[0]), 'name': row[1] or '', 'crop': row[2] or '',
                'tank_a': TankConfig(capacity=row[5] or 1000, dilution=row[6] or 100, dose=row[11] or 10),
                'tank_b': TankConfig(capacity=row[7] or 1000, dilution=row[8] or 100, dose=row[11] or 10),
                'tank_c': TankConfig(capacity=row[9] or 1000, dilution=row[10] or 100, dose=row[11] or 10),
                'micro': TankConfig(capacity=row[7] or 1000, dilution=row[8] or 100, dose=row[11] or 10),
                'desired_ec': row[12] or 0,
                'mode': mode or 'standard',
                'drain_percentage': drain_pct,
                'drain_analysis': drain,
                'target_values': targets,
            }
            break

    if not recipe:
        wb.close()
        return None

    # Load fertilizer entries
    ws2 = wb[SHEET_RECIPE_FERTS]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] and int(row[0]) == recipe_id:
            fid = row[1]
            qty = row[2] or 0
            from ..engine.fertilizers import FERT_BY_ID
            fdef = FERT_BY_ID.get(fid)
            if fdef:
                tank_key = {'A': 'tank_a', 'B': 'tank_b', 'C': 'tank_c', 'micro': 'micro'}.get(fdef.tank)
                if tank_key:
                    recipe[tank_key].entries[fid] = float(qty)

    # Load water analysis (with Na support)
    ws3 = wb[SHEET_WATER]
    water = WaterAnalysis()
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if row[0] and int(row[0]) == recipe_id:
            # Old format: EC, NO3, K, Mg, Ca, P2O5, S, Cl, Fe, Mn, Zn, B, Cu, Mo (no Na)
            # New format: EC, NO3, K, Mg, Ca, P2O5, S, Cl, Na, Fe, Mn, Zn, B, Cu, Mo
            fields_old = ['ec', 'no3', 'k', 'mg', 'ca', 'p2o5', 's', 'cl', 'fe', 'mn', 'zn', 'b', 'cu', 'mo']
            fields_new = ['ec', 'no3', 'k', 'mg', 'ca', 'p2o5', 's', 'cl', 'na', 'fe', 'mn', 'zn', 'b', 'cu', 'mo']
            # Detect format by column count
            cols = list(row[1:])  # skip recipe_id
            if len(cols) >= 15:
                fields = fields_new
            else:
                fields = fields_old
            for i, f in enumerate(fields):
                if i < len(cols):
                    setattr(water, f, float(cols[i] or 0))
            break
    recipe['water'] = water

    wb.close()
    return recipe


def save_recipe(
    recipe_id: Optional[int],
    name: str,
    crop: str,
    tank_a: TankConfig,
    tank_b: TankConfig,
    tank_c: TankConfig,
    micro: TankConfig,
    desired_ec: float,
    water: WaterAnalysis,
    mode: str = 'standard',
    drain_percentage: float = 0.0,
    drain_analysis: Optional[DrainAnalysis] = None,
    target_values: Optional[TargetValues] = None,
) -> int:
    path = ensure_database()
    wb = load_workbook(path)

    ws = wb[SHEET_RECIPES]
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    drain_json = json.dumps(drain_analysis.__dict__) if drain_analysis else '{}'
    target_json = json.dumps(target_values.__dict__) if target_values else '{}'

    if recipe_id is None:
        recipe_id = _next_recipe_id(ws)
        ws.append([
            recipe_id, name, crop, now, now,
            tank_a.capacity, tank_a.dilution,
            tank_b.capacity, tank_b.dilution,
            tank_c.capacity, tank_c.dilution,
            tank_a.dose, desired_ec,
            mode, drain_percentage, drain_json, target_json,
        ])
    else:
        for row in ws.iter_rows(min_row=2):
            if row[0].value and int(row[0].value) == recipe_id:
                row[1].value = name
                row[2].value = crop
                row[4].value = now
                row[5].value = tank_a.capacity
                row[6].value = tank_a.dilution
                row[7].value = tank_b.capacity
                row[8].value = tank_b.dilution
                row[9].value = tank_c.capacity
                row[10].value = tank_c.dilution
                row[11].value = tank_a.dose
                row[12].value = desired_ec
                # Extend row if needed
                for i in range(len(row), 17):
                    ws.cell(row=row[0].row, column=i + 1, value=None)
                ws.cell(row=row[0].row, column=14, value=mode)
                ws.cell(row=row[0].row, column=15, value=drain_percentage)
                ws.cell(row=row[0].row, column=16, value=drain_json)
                ws.cell(row=row[0].row, column=17, value=target_json)
                break

    # Clear old fertilizer entries
    ws2 = wb[SHEET_RECIPE_FERTS]
    rows_to_delete = []
    for i, row in enumerate(ws2.iter_rows(min_row=2), start=2):
        if row[0].value and int(row[0].value) == recipe_id:
            rows_to_delete.append(i)
    for i in reversed(rows_to_delete):
        ws2.delete_rows(i)

    for tank_key, tank in [('tank_a', tank_a), ('tank_b', tank_b), ('tank_c', tank_c), ('micro', micro)]:
        for fid, qty in tank.entries.items():
            if qty > 0:
                ws2.append([recipe_id, fid, qty])

    # Update water analysis (with Na)
    ws3 = wb[SHEET_WATER]
    found_water = False
    for row in ws3.iter_rows(min_row=2):
        if row[0].value and int(row[0].value) == recipe_id:
            vals = [water.ec, water.no3, water.k, water.mg, water.ca,
                    water.p2o5, water.s, water.cl, water.na,
                    water.fe, water.mn, water.zn, water.b, water.cu, water.mo]
            for i, v in enumerate(vals):
                ws3.cell(row=row[0].row, column=i + 2, value=v)
            found_water = True
            break
    if not found_water:
        ws3.append([recipe_id, water.ec, water.no3, water.k, water.mg, water.ca,
                     water.p2o5, water.s, water.cl, water.na,
                     water.fe, water.mn, water.zn, water.b, water.cu, water.mo])

    wb.save(path)
    wb.close()
    return recipe_id


def delete_recipe(recipe_id: int):
    path = ensure_database()
    wb = load_workbook(path)

    for sheet_name in [SHEET_RECIPES, SHEET_RECIPE_FERTS, SHEET_WATER, SHEET_HISTORY]:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value and int(row[0].value) == recipe_id:
                rows.append(i)
        for i in reversed(rows):
            ws.delete_rows(i)

    wb.save(path)
    wb.close()
